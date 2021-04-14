import signal
import sys
import time
import json
from requests import get
from datetime import date, datetime
from openpyxl import Workbook


def get_conf() -> tuple:
    file_conf = open('conf.json', 'r')
    file_rout = open('routes.json', 'r')

    conf = json.load(file_conf)
    rout = json.load(file_rout)

    file_conf.close()
    file_rout.close()

    domain = conf.pop('domain')
    routes = rout.pop('routes_id')

    return domain, conf, routes


def get_price(url: str, params: dict) -> list:
    now = datetime.now()
    timestamp = f'{now.hour}:{now.minute}'
    try:
        response = get(url, params=params)
        content = response.json()
        data_list = [int(content.get('time')),
                     int(content.get('options')[0].get('min_price')),
                     int(content.get('options')[0].get('price')),
                     timestamp]
        return data_list
    except Exception as e:
        print(e)
        return [0, 0, 0, timestamp]


def write_data(book: Workbook(), row: int, data: list) -> None:
    book.append(data)


def crete_sheets(routes: dict) -> tuple:
    wb = Workbook()
    return wb, [wb.create_sheet(f'id{num}') for num, route in routes.items() if route['enable']]


def main_function():
    day = date.today()
    now = datetime.now()
    timestamp_start = f'{now.hour}-{now.minute}'

    url, params, routes = get_conf()

    wb, wss = crete_sheets(routes)

    def signal_handler(*args):
        now = datetime.now()
        timestamp_end = f'{now.hour}-{now.minute}'
        wb.save(f"data/{str(date.today())} data({timestamp_start} {timestamp_end}).xlsx")
        sys.exit()

    signal.signal(signal.SIGTERM, signal_handler)

    row = 1
    while True:
        if day == date.today():
            for num, route in routes.items():
                start_point = f"{route.get('first_point')[0]},{route.get('first_point')[1]}"
                end_point = f"{route.get('end_point')[0]},{route.get('end_point')[1]}"

                params['rll'] = f"{start_point}~{end_point}"
                result = get_price(url, params)
                params['rll'] = f"{end_point}~{start_point}"
                result.extend(get_price(url, params))

                print(result)

                write_data(wss[int(num)], row, result)

                row += 1
                time.sleep(60)

        else:
            now = datetime.now()
            wb.save(f"data/{str(day)} data({timestamp_start} {now.hour}-{now.minute}).xlsx")

            url, params, routes = get_conf()
            wb, wss = crete_sheets(routes)

            day = date.today()
            now = datetime.now()
            timestamp_start = f'{now.hour}-{now.minute}'


if __name__ == '__main__':
    main_function()
